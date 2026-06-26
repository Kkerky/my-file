from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SRC = Path(
    r"C:\ServiceNow01\outputs\mecm_security_mapping"
    r"\MECM读取信息到ServiceNow表字段映射_23项安全指摘_字段落表矩阵追加版624_ACC比較列追加版625_O-X判定修正版625_取得元追加版625_中文説明追加版625.xlsx"
)
OUT = SRC.with_name(SRC.stem + "_MECMカスタマイズ要否追加版626.xlsx")
SHEET = "12_フィールド格納マトリクス_JP"
TARGET = "レジストリ直接参照ではない。"


def judge(ws, row):
    source_note = str(ws.cell(row, 14).value or "")
    if TARGET not in source_note:
        return "", None

    mecm_source = str(ws.cell(row, 6).value or "")
    info_name = str(ws.cell(row, 4).value or "")
    mecm_attr = str(ws.cell(row, 5).value or "")

    if "MECM クライアント別インベントリ View" in mecm_source or "ResourceID" in mecm_attr:
        return "不要：SG-SCCM標準処理で端末CIへ関連付け。MECM側の新規カスタムは不要。", "no"

    if "v_GS_OPERATING_SYSTEM" in mecm_source:
        return "不要：MECM標準Hardware InventoryのOS情報。MECM側の新規カスタムは不要。", "no"

    if "v_GS_WORKSTATION_STATUS" in mecm_source:
        return "不要：MECM標準のInventory実行状態。MECM側の新規カスタムは不要。", "no"

    software_update_views = [
        "v_UpdateInfo",
        "v_UpdateCIs",
        "v_Update_ComplianceStatus",
        "v_UpdateState_Combined",
        "v_UpdateScanStatus",
    ]
    if any(view in mecm_source for view in software_update_views):
        return "不要：Software Updates／WSUS同期・Scan設定が前提。新規カスタムWMI／ETLは不要。", "no"

    if "v_GS_SYSTEM_ACCOUNT" in mecm_source:
        if "算出項目" in mecm_source or "Guest 有効判定" in info_name:
            return "条件付き不要：UserAccount取得自体は既存Inventoryで対応可能。Guest判定はレポート／ServiceNow側の算出。", "conditional"
        return "条件付き不要：既存Hardware Inventoryクラスの有効化確認は必要。新規WMI Class追加は不要。", "conditional"

    if "Guest 有効判定" in info_name:
        return "条件付き不要：UserAccount取得自体は既存Inventoryで対応可能。Guest判定はレポート／ServiceNow側の算出。", "conditional"

    if "v_GS_SERVICE" in mecm_source:
        return "条件付き不要：既存Service Inventoryクラスの有効化確認は必要。新規WMI Class追加は不要。", "conditional"

    endpoint_views = ["v_GS_AntimalwareHealthStatus", "v_EndpointProtectionStatus"]
    if any(view in mecm_source for view in endpoint_views):
        if "製品固有" in mecm_source or "Compliance Baseline" in mecm_source:
            return (
                "要確認：Defender標準管理なら既存Viewで足りる可能性あり。"
                "他社AV／製品固有項目はCompliance Baseline等が必要な場合あり。"
            ), "check"
        if "算出項目" in mecm_source:
            return "条件付き不要：基礎データはEndpoint Protection／Defender状態Viewで取得。期限判定などは別途算出。", "conditional"
        return "条件付き不要：Endpoint Protection／Defender状態収集が有効なら取得可能。新規MECMカスタムは不要。", "conditional"

    disk_views = [
        "v_GS_ENCRYPTABLE_VOLUME",
        "v_GS_PROTECTED_VOLUME_INFO",
        "v_GS_TPM",
        "v_GS_DISK",
    ]
    if any(view in mecm_source for view in disk_views):
        if any(token in info_name for token in ["BitLocker", "暗号化", "TPM"]):
            return "条件付き不要：BitLocker／TPM関連InventoryまたはBitLocker管理が有効な前提。新規WMI Class追加は通常不要。", "conditional"
        return "不要：Disk系の標準Inventory情報。MECM側の新規カスタムは不要。", "no"

    return "要確認：標準View／既存Inventoryで不足する場合はMECM側の追加設定またはカスタムが必要。", "check"


def main():
    wb = load_workbook(SRC)
    ws = wb[SHEET]
    col = 16

    for row in range(1, ws.max_row + 1):
        source_cell = ws.cell(row, 15)
        target_cell = ws.cell(row, col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)

    ws.cell(1, col).font = Font(name="Meiryo UI", size=9, bold=True, color="1F4E5F")
    ws.cell(1, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(2, col).value = "判定口径：不要＝新規WMI Class／configuration.mof／独自ETL不要。条件付き不要＝既存機能の有効化・設定確認は必要。"
    ws.cell(2, col).font = Font(name="Meiryo UI", size=9, color="666666")
    ws.cell(2, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(3, col).value = "MECM人工カスタマイズ要否\n（標準機能／既存設定で足りるか）"
    ws.cell(3, col).font = Font(name="Meiryo UI", bold=True, color="FFFFFF")
    ws.cell(3, col).fill = PatternFill("solid", fgColor="1F4E5F")
    ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    fills = {
        "no": PatternFill("solid", fgColor="D9EAD3"),
        "conditional": PatternFill("solid", fgColor="FFF2CC"),
        "check": PatternFill("solid", fgColor="FCE4D6"),
        "blank": PatternFill("solid", fgColor="F2F2F2"),
    }
    thin = Side(style="thin", color="D9E2EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    counts = {}
    matched = 0
    for row in range(4, ws.max_row + 1):
        value, category = judge(ws, row)
        cell = ws.cell(row, col)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        cell.font = Font(name="Meiryo UI", size=9)
        cell.fill = fills.get(category or "blank")

        if value:
            matched += 1
            key = value.split("：", 1)[0]
            counts[key] = counts.get(key, 0) + 1

    ws.cell(1, col).value = f"対象 {matched} 行：" + "、".join(f"{key} {value} 行" for key, value in counts.items())
    ws.column_dimensions[get_column_letter(col)].width = 42
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 48

    try:
        for table in ws.tables.values():
            start = table.ref.split(":")[0]
            table.ref = f"{start}:{get_column_letter(col)}{ws.max_row}"
    except Exception:
        pass

    wb.save(OUT)
    print(OUT)
    print(f"matched={matched}")
    print(counts)


if __name__ == "__main__":
    main()
