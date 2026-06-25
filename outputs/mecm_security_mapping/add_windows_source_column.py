from copy import copy
from pathlib import Path
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


SOURCE = Path(os.environ["TARGET_XLSX"])
OUTPUT = SOURCE.with_name(SOURCE.stem + "_取得元追加版625.xlsx")


def registry(path: str, values: str, note: str = "") -> str:
    text = f"レジストリ：{path}\n値：{values}"
    if note:
        text += f"\n補足：{note}"
    return text


def wmi(namespace: str, class_name: str, properties: str, registry_note: str = "") -> str:
    text = (
        "レジストリ直接参照ではない。\n"
        f"WMI/CIM：{namespace}:{class_name}\n属性：{properties}"
    )
    if registry_note:
        text += f"\nレジストリ参考：{registry_note}"
    return text


def mecm(source: str, note: str = "") -> str:
    text = f"レジストリ直接参照ではない。\nMECM/WSUS DB：{source}"
    if note:
        text += f"\n補足：{note}"
    return text


def api(source: str, fields: str, note: str = "") -> str:
    text = f"レジストリ直接参照ではない。\nAPI/WMI：{source}\n属性：{fields}"
    if note:
        text += f"\n補足：{note}"
    return text


def external(source: str) -> str:
    return (
        "端末レジストリからは取得しない。\n"
        f"取得元：{source}"
    )


OS_CURRENT_VERSION = (
    r"HKLM\SOFTWARE\Microsoft\Windows NT\CurrentVersion"
)
UNINSTALL = (
    r"HKLM\SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall\*"
    "\n"
    r"HKLM\SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\*"
)
WINLOGON = r"HKLM\SOFTWARE\Microsoft\Windows NT\CurrentVersion\Winlogon"
SCREENSAVER_USER = (
    r"HKU\<User SID>\Control Panel\Desktop"
    "\n"
    r"HKU\<User SID>\Software\Policies\Microsoft\Windows\Control Panel\Desktop"
)
WINDOWS_UPDATE_AU = (
    r"HKLM\SOFTWARE\Policies\Microsoft\Windows\WindowsUpdate\AU"
)
DEFENDER_POLICY = (
    r"HKLM\SOFTWARE\Policies\Microsoft\Windows Defender"
)


SOURCES = {
    "OS 名": wmi(
        r"root\cimv2",
        "Win32_OperatingSystem",
        "Caption",
        OS_CURRENT_VERSION + r"\ProductName（参考値。MECM標準はWMI値を使用）",
    ),
    "OS バージョン": wmi(
        r"root\cimv2",
        "Win32_OperatingSystem",
        "Version",
        OS_CURRENT_VERSION + r"\DisplayVersion / CurrentVersion（参考値）",
    ),
    "OS Service Pack": wmi(
        r"root\cimv2",
        "Win32_OperatingSystem",
        "CSDVersion, ServicePackMajorVersion, ServicePackMinorVersion",
    ),
    "OS Build Number": wmi(
        r"root\cimv2",
        "Win32_OperatingSystem",
        "BuildNumber",
        OS_CURRENT_VERSION + r"\CurrentBuild / CurrentBuildNumber / UBR",
    ),
    "最終ハードウェアインベントリ / 検出日時": mecm(
        "v_GS_WORKSTATION_STATUS.LastHWScan",
        "MECMクライアントのHardware Inventory実行結果。端末の単一レジストリ値ではない。",
    ),
    "サポート終了日": external("Microsoft製品ライフサイクル情報または顧客ライフサイクルマスタ"),
    "ライフサイクル段階 / EOL 判定結果": external(
        "OS名・バージョンとライフサイクルマスタの照合・計算"
    ),
    "移行期限日": external("顧客管理の移行期限マスタ"),
    "期限までの残日数": external("期限日－判定日の計算結果"),
    "3 か月前警告判定": external("顧客定義の期限・警告ルールによる計算結果"),
    "ライフサイクル段階": external("顧客ライフサイクルマスタ"),
    "移行期間判定結果": external("OS/製品情報と顧客ライフサイクルマスタの照合結果"),
    "KB / Article ID": mecm(
        "v_UpdateInfo.ArticleID / v_UpdateCIs",
        "更新カタログおよびMECM/WSUS同期データ。端末レジストリではない。",
    ),
    "パッチタイトル": mecm("v_UpdateInfo.Title", "MECM/WSUS更新カタログ。"),
    "パッチ公開日": mecm(
        "v_UpdateInfo.DateCreated / DatePosted",
        "MECM/WSUS更新カタログ。",
    ),
    "パッチ重要度": mecm(
        "v_UpdateInfo.Severity / BulletinSeverity",
        "MECM/WSUS更新カタログ。",
    ),
    "端末のパッチ準拠 / 検出状態": mecm(
        "v_Update_ComplianceStatus / v_UpdateState_Combined",
        "MECMクライアントのUpdate Scan結果とState Messageを集約。",
    ),
    "最終ステータス確認日時": mecm(
        "v_Update_ComplianceStatus.LastStatusCheckTime",
        "MECM Site DBの状態時刻。",
    ),
    "最終更新スキャン日時": mecm(
        "v_UpdateScanStatus.LastScanTime",
        "Windows Update Agent/MECM Scan Agentの実行結果。",
    ),
    "未適用日数 / 週数および最終判定": external(
        "パッチ公開日、準拠状態および顧客監査ルールから算出"
    ),
    "アカウント名": wmi(
        r"root\cimv2", "Win32_UserAccount", "Name"
    ),
    "アカウント SID": wmi(
        r"root\cimv2", "Win32_UserAccount", "SID"
    ),
    "アカウントの有効 / 無効": wmi(
        r"root\cimv2", "Win32_UserAccount", "Disabled"
    ),
    "ローカル / ドメインアカウント属性": wmi(
        r"root\cimv2", "Win32_UserAccount", "LocalAccount, Domain"
    ),
    "Guest 有効判定": wmi(
        r"root\cimv2",
        "Win32_UserAccount",
        "SID末尾-501（Built-in Guest）とDisabledを組み合わせて判定",
        "SAMレジストリは保護されているため直接参照しない。",
    ),
    "AutoAdminLogon 値": registry(
        WINLOGON, "AutoAdminLogon"
    ),
    "DefaultUserName": registry(
        WINLOGON, "DefaultUserName（必要に応じてDefaultDomainName）"
    ),
    "レジストリパス": registry(
        WINLOGON,
        "AutoAdminLogon, DefaultUserName, DefaultDomainName",
        "DefaultPasswordは機密情報のため収集・格納対象にしない。",
    ),
    "自動ログオン有効判定": registry(
        WINLOGON,
        "AutoAdminLogon=1 およびDefaultUserName等の組み合わせ",
        "判定ロジックが必要。",
    ),
    "スクリーンセーバーの有効 / 無効": registry(
        SCREENSAVER_USER,
        "ScreenSaveActive, SCRNSAVE.EXE",
        "ユーザー単位のためHKU\\<SID>を確認する。",
    ),
    "パスワード保護の有効 / 無効": registry(
        SCREENSAVER_USER, "ScreenSaverIsSecure"
    ),
    "スクリーンセーバーのタイムアウト": registry(
        SCREENSAVER_USER, "ScreenSaveTimeOut"
    ),
    "ポリシー / レジストリパス": registry(
        SCREENSAVER_USER,
        "ScreenSaveActive, ScreenSaverIsSecure, ScreenSaveTimeOut, SCRNSAVE.EXE",
    ),
    "スクリーンセーバーのパスワード保護無効判定": registry(
        SCREENSAVER_USER,
        "ScreenSaveActive / ScreenSaverIsSecure / ScreenSaveTimeOut",
        "複数値を組み合わせて判定。",
    ),
    "Windows Update 自動更新ポリシー": registry(
        WINDOWS_UPDATE_AU, "NoAutoUpdate, AUOptions"
    ),
    "最終更新スキャン状態": api(
        "Windows Update Agent / MECM Scan Agent",
        "Scan state / v_UpdateScanStatus",
        "単一の信頼できるレジストリ値ではない。",
    ),
    "最終スキャンエラーコード": api(
        "Windows Update Agent / MECM Scan Agent",
        "LastErrorCode / v_UpdateScanStatus",
    ),
    "Windows Update Agent バージョン": api(
        "Windows Update Agent",
        r"%SystemRoot%\System32\wuaueng.dll のFileVersion等",
        "レジストリ値ではなくファイル/API情報。",
    ),
    "自動更新無効判定": registry(
        WINDOWS_UPDATE_AU,
        "NoAutoUpdate, AUOptions",
        "ポリシー値を組み合わせて判定。",
    ),
    "Windows Service 名": wmi(
        r"root\cimv2",
        "Win32_Service",
        "Name",
        r"HKLM\SYSTEM\CurrentControlSet\Services\<ServiceName>（サブキー名）",
    ),
    "表示名": wmi(
        r"root\cimv2",
        "Win32_Service",
        "DisplayName",
        r"HKLM\SYSTEM\CurrentControlSet\Services\<ServiceName>\DisplayName",
    ),
    "実行状態": wmi(
        r"root\cimv2",
        "Win32_Service",
        "State, Status",
        "実行状態は動的情報のためレジストリからは取得しない。",
    ),
    "起動種類": wmi(
        r"root\cimv2",
        "Win32_Service",
        "StartMode",
        r"HKLM\SYSTEM\CurrentControlSet\Services\<ServiceName>\Start",
    ),
    "実行ファイルパス": wmi(
        r"root\cimv2",
        "Win32_Service",
        "PathName",
        r"HKLM\SYSTEM\CurrentControlSet\Services\<ServiceName>\ImagePath",
    ),
    "実行アカウント": wmi(
        r"root\cimv2",
        "Win32_Service",
        "StartName",
        r"HKLM\SYSTEM\CurrentControlSet\Services\<ServiceName>\ObjectName",
    ),
    "不要サービス判定": external(
        "Win32_ServiceのName/State/StartModeと顧客の許可・禁止サービスリストを照合"
    ),
    "AV 製品名": registry(
        UNINSTALL,
        "DisplayName",
        r"稼働中AVの判定はroot\SecurityCenter2:AntiVirusProductまたは製品APIを使用。",
    ),
    "AV 製品バージョン": registry(
        UNINSTALL, "DisplayVersion"
    ),
    "AV 製品 Publisher / ベンダー": registry(
        UNINSTALL, "Publisher"
    ),
    "AV 製品インストール日": registry(
        UNINSTALL,
        "InstallDate",
        "製品によって値が存在しない、または更新日として書き換わる場合がある。",
    ),
    "インストール先 CI": mecm(
        "MECM ResourceIDと端末CIの関連付け",
        "レジストリ値ではなくConnector/IREによる端末関連付け。",
    ),
    "AV / Antimalware の有効 / 無効": api(
        r"root\SecurityCenter2:AntiVirusProduct / Microsoft Defender Get-MpComputerStatus",
        "productState / AntivirusEnabled / AMRunningMode",
        "Windows Serverや他社製品では利用可能なProviderが異なる。",
    ),
    "Endpoint Protection の保護状態": api(
        "MECM Endpoint Protection State / Microsoft Defender Get-MpComputerStatus",
        "RealTimeProtectionEnabled, AntivirusEnabled, AMServiceEnabled等",
    ),
    "指定 AV 製品との照合結果": external(
        "AV製品名・Publisherと顧客指定AV製品マスタの照合結果"
    ),
    "AV 製品サポート状態": external(
        "AV製品名・バージョンと製品ライフサイクルマスタの照合"
    ),
    "未導入 / 指定外 / サポート終了判定": external(
        "AVインベントリ、指定製品マスタ、製品ライフサイクルマスタの照合"
    ),
    "残日数": external("製品サポート終了日－判定日の計算結果"),
    "移行段階判定結果": external(
        "製品ライフサイクルマスタと判定日の照合結果"
    ),
    "AutoProtect / リアルタイム保護の有効 / 無効": api(
        "Microsoft Defender Get-MpComputerStatus / 他社AV製品API",
        "RealTimeProtectionEnabled等",
        DEFENDER_POLICY
        + r"\Real-Time Protection\DisableRealtimeMonitoring はポリシー参考値であり、実効状態はAPIで確認。",
    ),
    "AV ポリシー適用状態": api(
        "Microsoft Defender PowerShell/WMIまたは各AV製品の管理API",
        "製品固有の適用済みポリシー状態",
        "Defenderの設定候補：" + DEFENDER_POLICY,
    ),
    "AutoProtect 未設定判定": external(
        "リアルタイム保護状態と顧客判定ルールの組み合わせ"
    ),
    "定義ファイル / セキュリティインテリジェンスバージョン": api(
        "Microsoft Defender Get-MpComputerStatus / MSFT_MpComputerStatus",
        "AntivirusSignatureVersion",
    ),
    "定義ファイル最終更新日時": api(
        "Microsoft Defender Get-MpComputerStatus / MSFT_MpComputerStatus",
        "AntivirusSignatureLastUpdated",
    ),
    "Antimalware Engine バージョン": api(
        "Microsoft Defender Get-MpComputerStatus / MSFT_MpComputerStatus",
        "AMEngineVersion",
    ),
    "AV の有効 / 無効": api(
        "Microsoft Defender Get-MpComputerStatus / AntiVirusProduct",
        "AntivirusEnabled / productState",
    ),
    "定義ファイル更新後の経過日数": external(
        "AntivirusSignatureLastUpdatedと判定日の差分計算"
    ),
    "1 週間以上未更新の判定": external(
        "定義ファイル最終更新日時と顧客判定ルールの組み合わせ"
    ),
    "最終スキャン日時": api(
        "Microsoft Defender Get-MpComputerStatus / MSFT_MpComputerStatus",
        "QuickScanEndTime, FullScanEndTime",
    ),
    "スキャン種別": api(
        "Microsoft Defender Get-MpComputerStatus / Defenderイベント",
        "Quick Scan / Full Scan",
    ),
    "スキャン結果 / ヘルス状態": api(
        "Microsoft Defender PowerShell/WMI、イベントログまたはMECM Endpoint Protection",
        "Scan result / health state",
    ),
    "最終スキャンからの経過日数": external(
        "QuickScanEndTime/FullScanEndTimeと判定日の差分計算"
    ),
    "1 週間以上未スキャンの判定": external(
        "最終スキャン日時と顧客判定ルールの組み合わせ"
    ),
    "ディスク / ボリューム名": wmi(
        r"root\cimv2",
        "Win32_LogicalDisk / Win32_Volume",
        "DeviceID, Name, DriveLetter",
    ),
    "ディスク容量": wmi(
        r"root\cimv2",
        "Win32_LogicalDisk / Win32_DiskDrive",
        "Size, FreeSpace",
    ),
    "ドライブ種類": wmi(
        r"root\cimv2", "Win32_LogicalDisk", "DriveType"
    ),
    "BitLocker 保護状態": api(
        r"root\CIMV2\Security\MicrosoftVolumeEncryption:Win32_EncryptableVolume",
        "GetProtectionStatus / ProtectionStatus",
    ),
    "暗号化変換状態": api(
        r"root\CIMV2\Security\MicrosoftVolumeEncryption:Win32_EncryptableVolume",
        "GetConversionStatus / ConversionStatus, EncryptionPercentage",
    ),
    "暗号化方式": api(
        r"root\CIMV2\Security\MicrosoftVolumeEncryption:Win32_EncryptableVolume",
        "GetEncryptionMethod / EncryptionMethod",
    ),
    "TPM 状態 / バージョン": api(
        r"root\CIMV2\Security\MicrosoftTpm:Win32_Tpm",
        "IsEnabled_InitialValue, IsActivated_InitialValue, SpecVersion",
    ),
    "HDD 未暗号化判定": external(
        "Win32_EncryptableVolumeのProtectionStatus/ConversionStatusと顧客判定ルール"
    ),
    "ソフトウェア名": registry(
        UNINSTALL, "DisplayName"
    ),
    "ソフトウェアバージョン": registry(
        UNINSTALL, "DisplayVersion"
    ),
    "ソフトウェア Publisher / ベンダー": registry(
        UNINSTALL, "Publisher"
    ),
    "ソフトウェアインストール日": registry(
        UNINSTALL,
        "InstallDate",
        "全製品に存在するとは限らず、更新で変更される場合がある。",
    ),
    "ソフトウェア分類 / Category": external(
        "MECM Asset Intelligence Catalog / Software Catalogによる分類"
    ),
    "禁止ソフトウェアリストとの照合結果": external(
        "ソフトウェア名・Publisherと顧客禁止ソフトウェアリストの照合"
    ),
    "リスクレベル": external(
        "禁止ソフトウェア照合結果と顧客リスクルールによる算出"
    ),
    "最終指摘判定": external(
        "取得したソフトウェア情報と顧客監査ルールによる算出"
    ),
    "製品ライフサイクル段階": external(
        "ソフトウェア名・バージョンと製品ライフサイクルマスタの照合"
    ),
    "サポート終了日 / 移行期限日": external(
        "顧客製品ライフサイクルマスタ"
    ),
    "ライフサイクル指摘判定": external(
        "ソフトウェア情報、期限日および顧客判定ルールによる算出"
    ),
}


REFERENCES = [
    "https://learn.microsoft.com/en-us/windows/win32/cimwin32prov/win32-operatingsystem",
    "https://learn.microsoft.com/en-us/windows/win32/cimwin32prov/win32-useraccount",
    "https://learn.microsoft.com/en-us/windows/win32/cimwin32prov/win32-service",
    "https://learn.microsoft.com/en-us/windows/win32/cimwin32prov/win32-logicaldisk",
    "https://learn.microsoft.com/en-us/windows/win32/secprov/win32-encryptablevolume",
    "https://learn.microsoft.com/en-us/windows/win32/msi/uninstall-registry-key",
    "https://learn.microsoft.com/en-us/troubleshoot/windows-server/user-profiles-and-logon/turn-on-automatic-logon",
    "https://learn.microsoft.com/en-us/windows/deployment/update/waas-wu-settings",
    "https://learn.microsoft.com/en-us/defender-endpoint/microsoft-defender-antivirus-using-powershell",
    "https://learn.microsoft.com/en-us/intune/configmgr/develop/core/understand/sqlviews/hardware-inventory-views-configuration-manager",
]


def main() -> None:
    workbook = load_workbook(SOURCE)
    sheet = next(
        ws
        for ws in workbook.worksheets
        if ws.title.endswith("_JP") and not ws.title.endswith("(2)")
    )

    merged = {str(item) for item in sheet.merged_cells.ranges}
    if "A1:M1" in merged:
        sheet.unmerge_cells("A1:M1")
        sheet.merge_cells("A1:N1")
    if "A189:M191" in merged:
        sheet.unmerge_cells("A189:M191")
        sheet.merge_cells("A189:N191")

    sheet["A1"] = (
        "MECM セキュリティ指摘 23 項目：取得フィールド、"
        "ServiceNow格納可否、ACC比較およびWindows取得元"
    )

    header = sheet["N3"]
    header.value = "取得元\n（レジストリ／WMI／MECM DB等）"
    header._style = copy(sheet["M3"]._style)
    header.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    missing = []
    for row in range(4, 185):
        item = str(sheet.cell(row=row, column=4).value or "")
        source = SOURCES.get(item)
        if not source:
            missing.append((row, item))
            source = "取得元要確認：顧客MECMのHardware Inventory Class / Compliance Baselineを確認。"
        cell = sheet.cell(row=row, column=14)
        cell.value = source
        cell._style = copy(sheet["M" + str(row)]._style)
        cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        if source.startswith("レジストリ："):
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
        elif source.startswith("レジストリ直接参照ではない。"):
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        else:
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
        sheet.row_dimensions[row].height = max(
            sheet.row_dimensions[row].height or 15, 82
        )

    sheet.column_dimensions["N"].width = 60
    sheet["A189"] = (
        str(sheet["A189"].value)
        + "\n\n取得元列の読み方：レジストリに正式な値がある項目だけレジストリパスと値名を記載した。"
        "MECM Hardware Inventoryは多くの項目をWMI/CIMクラスから収集するため、"
        "WMI、MECM/WSUS DB、Windows API、製品API、外部マスタまたは計算結果を"
        "レジストリ値として扱わない。実際に収集されるクラスと項目は、顧客環境の"
        "Client Settings > Hardware Inventory Classes、Configuration.mof、"
        "Compliance BaselineおよびResource Explorerで確認する。確認日：2026-06-25。\n"
        + "\n".join(REFERENCES)
    )
    sheet["A189"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    sheet.row_dimensions[189].height = 270

    workbook.save(OUTPUT)
    print(f"output={OUTPUT}")
    print(f"missing={missing}")


if __name__ == "__main__":
    main()
