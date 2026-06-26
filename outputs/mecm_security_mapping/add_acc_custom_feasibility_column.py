from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\ServiceNow01\outputs\mecm_security_mapping")
SRC = sorted(
    BASE.glob("*I列X補足説明追加版626.xlsx"),
    key=lambda path: path.stat().st_mtime,
    reverse=True,
)[0]
OUT = SRC.with_name(SRC.stem + "_ACCカスタマイズ現実性追加版626.xlsx")


def text(*values):
    return " ".join(str(value or "") for value in values)


def is_standard_acc(m_value):
    return "○" in m_value and "標準取得" in m_value and "標準格納" in m_value


def is_acc_negative(m_value):
    return (
        "×" in m_value
        or "カスタマイズなしでは取得" in m_value
        or "ACC単独では取得" in m_value
        or "ACC単独では" in m_value
    )


def classify(category, finding, info, mecm_source, acc_note):
    combined = text(category, finding, info, mecm_source, acc_note)

    if not info and not acc_note:
        return ("", "", "blank")

    if is_standard_acc(acc_note):
        return (
            "標準で対応可能",
            "ACC標準収集とServiceNow標準CMDB格納の対象として整理済み。追加開発前提ではなく、PoCでは通常機能で確認する範囲。",
            "standard",
        )

    if not is_acc_negative(acc_note):
        return (
            "要確認",
            "既存列だけではACC標準可否を明確に判定できないため、実機または公式対象項目で再確認が必要。",
            "check",
        )

    if "ACC単独では取得" in acc_note or any(
        token in combined
        for token in [
            "サポート終了日",
            "移行期限日",
            "期限までの残日数",
            "3 か月前警告判定",
            "ライフサイクル",
            "EOL",
            "パッチ公開日",
            "パッチ重要度",
            "未適用",
            "ComplianceStatus",
            "UpdateInfo",
            "UpdateCIs",
            "製品サポート状態",
            "サポート終了製品",
            "不適切ソフトウェア分類",
            "危険度",
            "不適切ソフト判定",
            "禁止ソフト",
            "顧客ライフサイクルマスタ",
            "顧客管理の移行期限マスタ",
        ]
    ):
        return (
            "C：ACC単独では困難／非推奨",
            "端末内の情報だけでは判定できない。MECM／WSUSの更新カタログ、製品ライフサイクルマスタ、禁止ソフト一覧、AV管理基盤など外部マスタとの照合が必要。顧客へは「ACCだけで簡単に実現できる範囲ではない」と説明する。",
            "hard",
        )

    if any(
        token in combined
        for token in [
            "OS Build Number",
            "OS Service Pack",
            "最終ハードウェアインベントリ",
            "ResourceID",
            "端末CI",
            "SID",
            "UserAccount",
            "LocalAccount",
        ]
    ):
        return (
            "A：比較的簡単",
            "ACCの標準CMDBマッピング対象ではないが、端末上のWMI／CIM、PowerShell、osquery等で値そのものは取得しやすい。ただし、ServiceNowへの格納はカスタム項目・カスタムテーブル・Flow/ETLなどの設計が必要。",
            "easy",
        )

    if any(
        token in combined
        for token in [
            "AutoAdminLogon",
            "ScreenSave",
            "スクリーンセーバー",
            "自動ログオン",
            "Windows Update",
            "AUOptions",
            "NoAutoUpdate",
            "Windows Service",
            "不要なサービス",
            "BitLocker",
            "TPM",
            "暗号化",
            "Antimalware",
            "Endpoint Protection",
            "Defender",
            "Virus",
            "AutoProtect",
            "定義ファイル",
            "定時スキャン",
        ]
    ):
        return (
            "B：可能だが要設計",
            "端末から取得できる可能性はあるが、OS差分、製品差分、権限、判定ロジック、保存先、運用ルールの設計が必要。顧客へは「PoCで個別検証が必要。標準機能だけで完結するとは言えない」と説明する。",
            "design",
        )

    return (
        "B：可能だが要設計",
        "技術的にはカスタム収集の余地があるが、標準機能として保証できない。取得方法、格納先、判定ルール、保守担当を先に決める必要がある。",
        "design",
    )


def copy_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.border:
        target.border = copy(source.border)
    if source.fill:
        target.fill = copy(source.fill)
    if source.font:
        target.font = copy(source.font)


def main():
    wb = load_workbook(SRC)
    sheet_name = next(name for name in wb.sheetnames if name.startswith("12_"))
    ws = wb[sheet_name]

    start_col = 18
    class_col = start_col
    reason_col = start_col + 1

    for row in range(1, ws.max_row + 1):
        copy_style(ws.cell(row, 17), ws.cell(row, class_col))
        copy_style(ws.cell(row, 17), ws.cell(row, reason_col))

    ws.cell(1, class_col).value = "ACCカスタマイズ現実性分類を追加。標準可、比較的簡単、要設計、困難／非推奨に分け、顧客説明で過大約束しないための補足。"
    ws.cell(1, reason_col).value = "判断理由と顧客説明用コメント。"
    for col in (class_col, reason_col):
        ws.cell(1, col).font = Font(name="Meiryo UI", size=9, bold=True, color="1F4E5F")
        ws.cell(1, col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(2, class_col).value = (
        "分類基準：標準で対応可能＝ACC標準収集・標準格納。"
        "A＝端末値の取得は比較的容易。B＝取得可能性はあるが設計・個別検証が必要。"
        "C＝ACC単独では困難、MECM/WSUS/外部マスタ等を使うべき。"
    )
    ws.cell(2, reason_col).value = "C判定は、理論上スクリプトで一部値を取れる場合でも、顧客へ「ACCだけで簡単にできる」と説明しない対象。"
    for col in (class_col, reason_col):
        ws.cell(2, col).font = Font(name="Meiryo UI", size=9, color="666666")
        ws.cell(2, col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(3, class_col).value = "ACCカスタマイズ現実性\n（簡単／要設計／困難）"
    ws.cell(3, reason_col).value = "判断理由・顧客説明用コメント"
    for col in (class_col, reason_col):
        ws.cell(3, col).font = Font(name="Meiryo UI", size=9, bold=True, color="FFFFFF")
        ws.cell(3, col).fill = PatternFill("solid", fgColor="1F4E5F")
        ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    fills = {
        "standard": PatternFill("solid", fgColor="D9EAD3"),
        "easy": PatternFill("solid", fgColor="E2F0D9"),
        "design": PatternFill("solid", fgColor="FFF2CC"),
        "hard": PatternFill("solid", fgColor="F4CCCC"),
        "check": PatternFill("solid", fgColor="FCE4D6"),
        "blank": PatternFill("solid", fgColor="F2F2F2"),
    }
    thin = Side(style="thin", color="D9E2EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    category = ""
    finding = ""
    counts = {}
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 2).value:
            category = str(ws.cell(row, 2).value)
        if ws.cell(row, 3).value:
            finding = str(ws.cell(row, 3).value)

        info = str(ws.cell(row, 4).value or "")
        mecm_source = str(ws.cell(row, 6).value or "")
        acc_note = str(ws.cell(row, 13).value or "")
        label, reason, key = classify(category, finding, info, mecm_source, acc_note)

        ws.cell(row, class_col).value = label
        ws.cell(row, reason_col).value = reason
        for col in (class_col, reason_col):
            cell = ws.cell(row, col)
            cell.font = Font(name="Meiryo UI", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.fill = fills[key]
        if label:
            counts[label] = counts.get(label, 0) + 1

    summary = "、".join(f"{key} {value}行" for key, value in counts.items())
    ws.cell(1, class_col).value = f"ACCカスタマイズ現実性分類を追加：{summary}"

    ws.column_dimensions[get_column_letter(class_col)].width = 28
    ws.column_dimensions[get_column_letter(reason_col)].width = 62
    ws.row_dimensions[2].height = 56
    ws.row_dimensions[3].height = 48

    try:
        for table in ws.tables.values():
            start = table.ref.split(":")[0]
            table.ref = f"{start}:{get_column_letter(reason_col)}{ws.max_row}"
    except Exception:
        pass

    wb.save(OUT)
    print(str(OUT))
    print(counts)


if __name__ == "__main__":
    main()
