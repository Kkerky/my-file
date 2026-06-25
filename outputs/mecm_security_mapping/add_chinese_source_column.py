from copy import copy
from pathlib import Path
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment


SOURCE = Path(os.environ["TARGET_XLSX"])
OUTPUT = SOURCE.with_name(SOURCE.stem + "_中文説明追加版625.xlsx")


REPLACEMENTS = [
    ("OS/製品情報と顧客ライフサイクルマスタの照合結果",
     "OS/产品信息与客户生命周期主数据的匹配结果"),
    ("はポリシー参考値であり、実効状態はAPIで確認。",
     "是策略参考值，实际生效状态需要通过API确认。"),
    ("はポリシー参考值であり、実効状態はAPIで確認。",
     "是策略参考值，实际生效状态需要通过API确认。"),
    ("レジストリ直接参照ではない。", "不是直接读取注册表。"),
    ("端末レジストリからは取得しない。", "不从终端注册表获取。"),
    ("レジストリ参考：", "注册表参考："),
    ("レジストリ：", "注册表："),
    ("取得元：", "数据来源："),
    ("属性：", "属性："),
    ("値：", "值："),
    ("補足：", "补充："),
    ("参考値。MECM標準はWMI値を使用", "仅供参考；MECM标准采集使用WMI值"),
    ("参考値", "参考值"),
    ("MECMクライアントのHardware Inventory実行結果。端末の単一レジストリ値ではない。",
     "MECM客户端执行Hardware Inventory后产生的结果，不是终端上的单一注册表值。"),
    ("Microsoft製品ライフサイクル情報または顧客ライフサイクルマスタ",
     "Microsoft产品生命周期信息或客户生命周期主数据"),
    ("OS名・バージョンとライフサイクルマスタの照合・計算",
     "将OS名称、版本与生命周期主数据进行匹配和计算"),
    ("顧客管理の移行期限マスタ", "客户维护的迁移期限主数据"),
    ("期限日－判定日の計算結果", "期限日期减去判定日期的计算结果"),
    ("顧客定義の期限・警告ルールによる計算結果", "按照客户定义的期限及警告规则计算的结果"),
    ("顧客ライフサイクルマスタ", "客户生命周期主数据"),
    ("OS/製品情報と顧客ライフサイクルマスタの照合結果",
     "OS/产品信息与客户生命周期主数据的匹配结果"),
    ("更新カタログおよびMECM/WSUS同期データ。端末レジストリではない。",
     "来自更新目录以及MECM/WSUS同步数据，不是终端注册表值。"),
    ("MECM/WSUS更新カタログ。", "MECM/WSUS更新目录。"),
    ("MECMクライアントのUpdate Scan結果とState Messageを集約。",
     "汇总MECM客户端的Update Scan结果和State Message。"),
    ("MECM Site DBの状態時刻。", "MECM Site DB中记录的状态时间。"),
    ("Windows Update Agent/MECM Scan Agentの実行結果。",
     "Windows Update Agent/MECM Scan Agent的执行结果。"),
    ("パッチ公開日、準拠状態および顧客監査ルールから算出",
     "根据补丁发布日期、合规状态和客户审计规则计算"),
    ("SID末尾-501（Built-in Guest）とDisabledを組み合わせて判定",
     "结合SID末尾-501（Built-in Guest）和Disabled进行判定"),
    ("SAMレジストリは保護されているため直接参照しない。",
     "SAM注册表受到保护，因此不直接读取。"),
    ("必要に応じてDefaultDomainName", "必要时同时参考DefaultDomainName"),
    ("DefaultPasswordは機密情報のため収集・格納対象にしない。",
     "DefaultPassword属于敏感信息，不应作为采集和存储对象。"),
    ("AutoAdminLogon=1 およびDefaultUserName等の組み合わせ",
     "组合判断AutoAdminLogon=1、DefaultUserName等值"),
    ("判定ロジックが必要。", "需要另外定义判定逻辑。"),
    ("ユーザー単位のためHKU\\<SID>を確認する。",
     "该设置按用户保存，需要检查HKU\\<SID>。"),
    ("複数値を組み合わせて判定。", "需要组合多个值进行判定。"),
    ("単一の信頼できるレジストリ値ではない。",
     "不存在能够单独代表该状态的可靠注册表值。"),
    ("%SystemRoot%\\System32\\wuaueng.dll のFileVersion等",
     "%SystemRoot%\\System32\\wuaueng.dll的FileVersion等"),
    ("レジストリ値ではなくファイル/API情報。",
     "信息来自文件/API，而不是注册表值。"),
    ("ポリシー値を組み合わせて判定。", "需要组合多个策略值进行判定。"),
    ("サブキー名", "子项名称"),
    ("実行状態は動的情報のためレジストリからは取得しない。",
     "运行状态属于动态信息，不从注册表读取。"),
    ("Win32_ServiceのName/State/StartModeと顧客の許可・禁止サービスリストを照合",
     "将Win32_Service的Name/State/StartMode与客户允许/禁止服务清单进行匹配"),
    ("稼働中AVの判定はroot\\SecurityCenter2:AntiVirusProductまたは製品APIを使用。",
     "判断正在运行的杀毒软件时，应使用root\\SecurityCenter2:AntiVirusProduct或产品API。"),
    ("製品によって値が存在しない、または更新日として書き換わる場合がある。",
     "部分产品可能不存在该值，或者软件更新后该值可能被改写。"),
    ("MECM ResourceIDと端末CIの関連付け", "MECM ResourceID与终端CI的关联"),
    ("レジストリ値ではなくConnector/IREによる端末関連付け。",
     "该信息不是注册表值，而是由Connector/IRE建立终端关联。"),
    ("Windows Serverや他社製品では利用可能なProviderが異なる。",
     "Windows Server及第三方产品可使用的Provider可能不同。"),
    ("RealTimeProtectionEnabled, AntivirusEnabled, AMServiceEnabled等",
     "RealTimeProtectionEnabled、AntivirusEnabled、AMServiceEnabled等"),
    ("AV製品名・Publisherと顧客指定AV製品マスタの照合結果",
     "杀毒软件名称、Publisher与客户指定杀毒产品主数据的匹配结果"),
    ("AV製品名・バージョンと製品ライフサイクルマスタの照合",
     "将杀毒软件名称、版本与产品生命周期主数据进行匹配"),
    ("AVインベントリ、指定製品マスタ、製品ライフサイクルマスタの照合",
     "对杀毒软件清单、指定产品主数据和产品生命周期主数据进行匹配"),
    ("製品サポート終了日－判定日の計算結果",
     "产品支持结束日期减去判定日期的计算结果"),
    ("製品ライフサイクルマスタと判定日の照合結果",
     "产品生命周期主数据与判定日期的匹配结果"),
    ("他社AV製品API", "第三方杀毒产品API"),
    ("はポリシー参考値であり、実効状態はAPIで確認。",
     "是策略参考值，实际生效状态需要通过API确认。"),
    ("または各AV製品の管理API", "或各杀毒产品的管理API"),
    ("製品固有の適用済みポリシー状態", "产品特有的已应用策略状态"),
    ("Defenderの設定候補：", "Defender设置参考："),
    ("リアルタイム保護状態と顧客判定ルールの組み合わせ",
     "组合实时保护状态和客户判定规则"),
    ("AntivirusSignatureLastUpdatedと判定日の差分計算",
     "计算AntivirusSignatureLastUpdated与判定日期之间的时间差"),
    ("定義ファイル最終更新日時と顧客判定ルールの組み合わせ",
     "组合病毒定义文件最后更新时间和客户判定规则"),
    ("Defenderイベント", "Defender事件"),
    ("イベントログまたはMECM Endpoint Protection",
     "事件日志或MECM Endpoint Protection"),
    ("QuickScanEndTime/FullScanEndTimeと判定日の差分計算",
     "计算QuickScanEndTime/FullScanEndTime与判定日期之间的时间差"),
    ("最終スキャン日時と顧客判定ルールの組み合わせ",
     "组合最后扫描时间和客户判定规则"),
    ("Win32_EncryptableVolumeのProtectionStatus/ConversionStatusと顧客判定ルール",
     "组合Win32_EncryptableVolume的ProtectionStatus/ConversionStatus和客户判定规则"),
    ("全製品に存在するとは限らず、更新で変更される場合がある。",
     "并非所有产品都存在该值，软件更新后也可能发生变化。"),
    ("MECM Asset Intelligence Catalog / Software Catalogによる分類",
     "由MECM Asset Intelligence Catalog / Software Catalog进行分类"),
    ("ソフトウェア名・Publisherと顧客禁止ソフトウェアリストの照合",
     "将软件名称、Publisher与客户禁止软件清单进行匹配"),
    ("禁止ソフトウェア照合結果と顧客リスクルールによる算出",
     "根据禁止软件匹配结果和客户风险规则计算"),
    ("取得したソフトウェア情報と顧客監査ルールによる算出",
     "根据取得的软件信息和客户审计规则计算"),
    ("ソフトウェア名・バージョンと製品ライフサイクルマスタの照合",
     "将软件名称、版本与产品生命周期主数据进行匹配"),
    ("顧客製品ライフサイクルマスタ", "客户产品生命周期主数据"),
    ("ソフトウェア情報、期限日および顧客判定ルールによる算出",
     "根据软件信息、期限日期和客户判定规则计算"),
]


def translate(text: str) -> str:
    translated = text
    for source, target in REPLACEMENTS:
        translated = translated.replace(source, target)
    return translated


def main() -> None:
    workbook = load_workbook(SOURCE)
    sheet = next(
        ws
        for ws in workbook.worksheets
        if ws.title.endswith("_JP") and not ws.title.endswith("(2)")
    )

    merged = {str(item) for item in sheet.merged_cells.ranges}
    if "A1:N1" in merged:
        sheet.unmerge_cells("A1:N1")
        sheet.merge_cells("A1:O1")
    if "A189:N191" in merged:
        sheet.unmerge_cells("A189:N191")
        sheet.merge_cells("A189:O191")

    sheet["A1"] = (
        "MECM セキュリティ指摘 23 項目：取得フィールド、"
        "ServiceNow格納可否、ACC比較、Windows取得元および中国語説明"
    )

    header = sheet["O3"]
    header.value = "取得元の中国語説明\n（注册表／WMI／MECM DB等）"
    header._style = copy(sheet["N3"]._style)
    header.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    for row in range(4, 185):
        source_cell = sheet.cell(row=row, column=14)
        target_cell = sheet.cell(row=row, column=15)
        target_cell.value = translate(str(source_cell.value or ""))
        target_cell._style = copy(source_cell._style)
        target_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

    sheet.column_dimensions["O"].width = 60
    sheet["A189"] = (
        str(sheet["A189"].value)
        + "\n\nO列为N列来源说明的中文对照版。注册表路径、WMI/CIM类名、"
        "MECM数据库视图、API名称和字段名保持原始技术标识不变。"
    )
    sheet["A189"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    sheet.row_dimensions[189].height = 290

    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
