from copy import copy
from pathlib import Path
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


SOURCE = Path(os.environ["TARGET_XLSX"])
OUTPUT = SOURCE.with_name(
    SOURCE.stem + "_ACC比較列追加版625.xlsx"
)

ACC_DATA_URL = (
    "https://www.servicenow.com/docs/r/it-operations-management/"
    "agent-client-collector/acc-data-collection.html"
)
ACC_V_REFERENCE_URL = (
    "https://www.servicenow.com/docs/r/it-operations-management/"
    "agent-client-collector/agent-client-collector-for-visibility-references.html"
)
ACC_V_POLICY_URL = (
    "https://www.servicenow.com/docs/r/it-operations-management/"
    "agent-client-collector/agent-client-collector-for-visibility-checks-and-policies.html"
)
ACC_CHECK_URL = (
    "https://www.servicenow.com/docs/r/it-operations-management/"
    "agent-client-collector/acc-api-check-def.html"
)


def standard(destination: str, note: str = "") -> str:
    text = f"○ 標準取得・標準格納\n格納先：{destination}"
    if note:
        text += f"\n{note}"
    return text


def custom(method: str, destination: str, note: str = "") -> str:
    text = (
        f"△ カスタム取得・変換が必要\n取得：{method}\n"
        f"格納：{destination}"
    )
    if note:
        text += f"\n{note}"
    return text


def unavailable(reason: str, destination: str = "カスタム監査／Complianceテーブル") -> str:
    return (
        "× ACC単独では取得・判定不可\n"
        f"理由：{reason}\n格納：{destination}"
    )


STANDARD = {
    "OS 名": standard("cmdb_ci_computer.os", "ACC-F基本インベントリの公式対象。"),
    "OS バージョン": standard(
        "cmdb_ci_computer.os_version", "ACC-F基本インベントリの公式対象。"
    ),
    "アカウント名": standard(
        "cmdb_os_user", "ACC-VCのLocal Userモジュールでローカルユーザーを格納。"
    ),
    "AV 製品名": standard(
        "cmdb_sam_sw_install（SAM有）／cmdb_software_instance（SAM無）",
        "ACC-VCのインストール済みソフトウェアとして取得。稼働中AVの判定は別途必要。",
    ),
    "AV 製品バージョン": standard(
        "cmdb_sam_sw_install.version等",
        "ACC-VCのソフトウェアインベントリとして取得。",
    ),
    "AV 製品 Publisher / ベンダー": standard(
        "cmdb_sam_sw_install／Software Discovery Model",
        "ACC-VCとSAMの構成・正規化結果に依存。",
    ),
    "インストール先 CI": standard(
        "cmdb_sam_sw_install.installed_on／cmdb_software_instance",
        "対象Computer CIへの関連付けを標準処理。",
    ),
    "ディスク / ボリューム名": standard(
        "cmdb_ci_file_system／cmdb_ci_disk／cmdb_ci_storage_device",
        "ACC-VCのFile Systems／Storage Devicesモジュール対象。",
    ),
    "ディスク容量": standard(
        "cmdb_ci_file_system／cmdb_ci_disk／cmdb_ci_storage_device",
        "標準モジュール対象。実フィールドはACC-VC payloadで確認。",
    ),
    "ドライブ種類": standard(
        "cmdb_ci_file_system／cmdb_ci_disk／cmdb_ci_storage_device",
        "標準モジュール対象。実フィールドはACC-VC payloadで確認。",
    ),
    "ソフトウェア名": standard(
        "cmdb_sam_sw_install（SAM有）／cmdb_software_instance（SAM無）",
        "ACC-VC Software Installed Policyの対象。",
    ),
    "ソフトウェアバージョン": standard(
        "cmdb_sam_sw_install.version等",
        "ACC-VC Software Installed Policyの対象。",
    ),
    "ソフトウェア Publisher / ベンダー": standard(
        "cmdb_sam_sw_install／Software Discovery Model",
        "ACC-VCとSAMの正規化結果に依存。",
    ),
}


CUSTOM = {
    "OS Service Pack": custom(
        "osquery／PowerShell／WMI",
        "cmdb_ci_computer.os_service_pack（カスタムマッピング）",
        "ACC標準収集項目としての明記なし。",
    ),
    "OS Build Number": custom(
        "osquery system_info／Windowsレジストリ",
        "カスタムCI属性または監査テーブル",
        "ACC標準CMDB項目への自動格納なし。",
    ),
    "最終ハードウェアインベントリ / 検出日時": custom(
        "ACC実行日時／収集履歴",
        "cmdb_ci.last_discovered候補（実機確認）",
        "MECM LastHWScanと同一意味ではない。",
    ),
    "KB / Article ID": custom(
        "osquery patches／PowerShell",
        "カスタムPatch／Complianceテーブル",
        "取得できるのは主に適用済みパッチ。未適用パッチ一覧にはならない。",
    ),
    "パッチタイトル": custom(
        "osquery patches／PowerShell",
        "カスタムPatchテーブル",
        "端末から取得できる情報は限定的で、MECM更新カタログと同等ではない。",
    ),
    "最終ステータス確認日時": custom(
        "カスタムCheckの実行日時",
        "カスタムPatch／Complianceテーブル",
        "MECMのLastStatusCheckTimeとは別の収集時刻。",
    ),
    "最終更新スキャン日時": custom(
        "PowerShell／Windows Update API",
        "カスタムPatch／Complianceテーブル",
        "ACC標準CMDB項目ではない。",
    ),
    "アカウント SID": custom(
        "ACC-VC Local User payload／osquery users",
        "cmdb_os_userの利用可能フィールドまたはカスタム項目",
        "SIDの標準フィールドマッピングは対象インスタンスで確認。",
    ),
    "アカウントの有効 / 無効": custom(
        "osquery users／PowerShell",
        "cmdb_os_userのカスタム項目またはComplianceテーブル",
        "ローカルユーザー一覧の標準格納と、無効状態の格納は分けて確認。",
    ),
    "ローカル / ドメインアカウント属性": custom(
        "ACC-VC Local User payload／osquery users",
        "cmdb_os_userまたはカスタム項目",
        "標準モジュールはローカルユーザーが中心。",
    ),
    "Guest 有効判定": custom(
        "osquery users／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
        "ユーザー情報取得後に顧客ルールで判定。",
    ),
    "AutoAdminLogon 値": custom(
        "osquery registry／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "DefaultUserName": custom(
        "osquery registry／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "レジストリパス": custom(
        "osquery registry",
        "カスタムSecurity Finding／Evidenceテーブル",
    ),
    "自動ログオン有効判定": custom(
        "osquery registry＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "スクリーンセーバーの有効 / 無効": custom(
        "osquery registry／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "パスワード保護の有効 / 無効": custom(
        "osquery registry／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "スクリーンセーバーのタイムアウト": custom(
        "osquery registry／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "ポリシー / レジストリパス": custom(
        "osquery registry",
        "カスタムSecurity Finding／Evidenceテーブル",
    ),
    "スクリーンセーバーのパスワード保護無効判定": custom(
        "osquery registry＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "Windows Update 自動更新ポリシー": custom(
        "osquery windows_security_center／registry／PowerShell",
        "カスタムSecurity Finding／Complianceテーブル",
        "ACC標準CMDB属性ではない。",
    ),
    "最終更新スキャン状態": custom(
        "PowerShell／Windows Update API",
        "カスタムPatch／Complianceテーブル",
    ),
    "最終スキャンエラーコード": custom(
        "PowerShell／Windows Update API",
        "カスタムPatch／Complianceテーブル",
    ),
    "Windows Update Agent バージョン": custom(
        "PowerShell／ファイル・レジストリ情報",
        "カスタムPatch／Complianceテーブル",
    ),
    "自動更新無効判定": custom(
        "windows_security_center／registry＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "Windows Service 名": custom(
        "osquery services／PowerShell",
        "カスタムService Inventoryテーブル",
        "ACC標準のcmdb_running_processは実行中プロセスであり、サービス台帳とは異なる。",
    ),
    "表示名": custom(
        "osquery services／PowerShell",
        "カスタムService Inventoryテーブル",
    ),
    "実行状態": custom(
        "osquery services／PowerShell",
        "カスタムService Inventory／Monitoringテーブル",
    ),
    "起動種類": custom(
        "osquery services／PowerShell",
        "カスタムService Inventory／Complianceテーブル",
    ),
    "実行ファイルパス": custom(
        "osquery services／processes",
        "カスタムService Inventoryテーブル",
        "実行中プロセスはcmdb_running_processにも格納可能。",
    ),
    "実行アカウント": custom(
        "osquery services／PowerShell",
        "カスタムService Inventoryテーブル",
    ),
    "不要サービス判定": custom(
        "サービス情報＋顧客許可リスト",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "AV 製品インストール日": custom(
        "ACC-VC Installed Software payload",
        "cmdb_sam_sw_install.install_date候補",
        "値の有無はOS・製品・ACC-VCバージョンに依存。",
    ),
    "AV / Antimalware の有効 / 無効": custom(
        "osquery windows_security_center／Security Center API",
        "カスタムSecurity Finding／Complianceテーブル",
        "ソフトウェア導入情報とは別の稼働状態。",
    ),
    "Endpoint Protection の保護状態": custom(
        "windows_security_center／PowerShell／製品API",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "指定 AV 製品との照合結果": custom(
        "ACC-VCソフトウェア情報＋顧客指定製品マスタ",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "未導入 / 指定外 / サポート終了判定": custom(
        "ACC-VCソフトウェア情報＋顧客製品・ライフサイクルマスタ",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "移行段階判定結果": custom(
        "ACC取得情報＋顧客ライフサイクルマスタ",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "AutoProtect / リアルタイム保護の有効 / 無効": custom(
        "windows_security_center／PowerShell／AV製品固有API",
        "カスタムSecurity Finding／Complianceテーブル",
        "製品依存のため標準Checkではない。",
    ),
    "AV ポリシー適用状態": custom(
        "PowerShell／AV製品固有API・レジストリ",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "AutoProtect 未設定判定": custom(
        "AV状態＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "定義ファイル / セキュリティインテリジェンスバージョン": custom(
        "PowerShell／AV製品固有API",
        "カスタムAV Health／Complianceテーブル",
    ),
    "定義ファイル最終更新日時": custom(
        "PowerShell／AV製品固有API",
        "カスタムAV Health／Complianceテーブル",
    ),
    "Antimalware Engine バージョン": custom(
        "PowerShell／AV製品固有API",
        "カスタムAV Health／Complianceテーブル",
    ),
    "AV の有効 / 無効": custom(
        "windows_security_center／PowerShell",
        "カスタムAV Health／Complianceテーブル",
    ),
    "定義ファイル更新後の経過日数": custom(
        "定義更新日時＋ServiceNow側計算",
        "カスタムAV Health／Complianceテーブル",
    ),
    "1 週間以上未更新の判定": custom(
        "定義更新日時＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "最終スキャン日時": custom(
        "PowerShell／AV製品固有API",
        "カスタムAV Health／Complianceテーブル",
    ),
    "スキャン種別": custom(
        "PowerShell／AV製品固有API",
        "カスタムAV Health／Complianceテーブル",
    ),
    "スキャン結果 / ヘルス状態": custom(
        "PowerShell／AV製品固有API",
        "カスタムAV Health／Complianceテーブル",
    ),
    "最終スキャンからの経過日数": custom(
        "最終スキャン日時＋ServiceNow側計算",
        "カスタムAV Health／Complianceテーブル",
    ),
    "1 週間以上未スキャンの判定": custom(
        "最終スキャン日時＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "BitLocker 保護状態": custom(
        "osquery bitlocker_info／PowerShell",
        "カスタムDisk Security／Complianceテーブル",
        "ACC-VC標準Storage Device属性には暗号化状態の明記なし。",
    ),
    "暗号化変換状態": custom(
        "osquery bitlocker_info／PowerShell",
        "カスタムDisk Security／Complianceテーブル",
    ),
    "暗号化方式": custom(
        "osquery bitlocker_info／PowerShell",
        "カスタムDisk Security／Complianceテーブル",
    ),
    "TPM 状態 / バージョン": custom(
        "osquery／PowerShell Get-Tpm",
        "カスタムDisk Security／Complianceテーブル",
    ),
    "HDD 未暗号化判定": custom(
        "BitLocker情報＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "ソフトウェアインストール日": custom(
        "ACC-VC Installed Software payload",
        "cmdb_sam_sw_install.install_date候補",
        "値の有無はOS・インストーラ・ACC-VCバージョンに依存。",
    ),
    "ソフトウェア分類 / Category": custom(
        "ACC-VCソフトウェア情報＋SAM正規化",
        "Software Discovery Model／SAM関連テーブル",
        "CategoryはACCの直接取得値ではない。",
    ),
    "禁止ソフトウェアリストとの照合結果": custom(
        "ACC-VCソフトウェア情報＋顧客禁止リスト",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "リスクレベル": custom(
        "ソフトウェア照合結果＋顧客リスクルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "最終指摘判定": custom(
        "ACC取得情報＋顧客判定ルール",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
    "ライフサイクル指摘判定": custom(
        "ACC-VCソフトウェア情報＋顧客ライフサイクルマスタ",
        "カスタムSecurity Finding／Complianceテーブル",
    ),
}


UNAVAILABLE = {
    "サポート終了日": unavailable(
        "OS／製品ライフサイクルマスタが必要で、端末インベントリ値ではない。"
    ),
    "ライフサイクル段階 / EOL 判定結果": unavailable(
        "外部ライフサイクルマスタとの照合結果。"
    ),
    "移行期限日": unavailable("顧客管理の移行期限マスタが必要。"),
    "期限までの残日数": unavailable("期限日が外部マスタ値であり、判定計算が必要。"),
    "3 か月前警告判定": unavailable("顧客定義の期限・警告ルールによる算出結果。"),
    "ライフサイクル段階": unavailable("顧客ライフサイクルマスタによる分類。"),
    "移行期間判定結果": unavailable("顧客ライフサイクルマスタとの照合結果。"),
    "パッチ公開日": unavailable(
        "端末情報だけでは完全な更新カタログの公開日を取得できない。MECM／Microsoft Updateカタログが必要。",
        "カスタムPatch Catalog／Complianceテーブル",
    ),
    "パッチ重要度": unavailable(
        "端末情報だけでは適用対象パッチの重要度カタログを保持しない。",
        "カスタムPatch Catalog／Complianceテーブル",
    ),
    "端末のパッチ準拠 / 検出状態": unavailable(
        "未適用を含む適用対象パッチ一覧と準拠評価が必要。ACCの適用済みパッチ照会だけでは不足。",
        "カスタムPatch Complianceテーブル",
    ),
    "未適用日数 / 週数および最終判定": unavailable(
        "パッチ公開日、適用対象判定および顧客監査ルールが必要。",
        "カスタムPatch Complianceテーブル",
    ),
    "AV 製品サポート状態": unavailable(
        "製品ライフサイクルマスタとの照合が必要。"
    ),
    "残日数": unavailable("外部ライフサイクル期限日からの算出値。"),
    "製品ライフサイクル段階": unavailable(
        "製品ライフサイクルマスタとの照合結果。"
    ),
    "サポート終了日 / 移行期限日": unavailable(
        "顧客製品ライフサイクルマスタの値。"
    ),
}


def assessment(value: str) -> str:
    if value in STANDARD:
        return STANDARD[value]
    if value in CUSTOM:
        return CUSTOM[value]
    if value in UNAVAILABLE:
        return UNAVAILABLE[value]
    return custom(
        "対象OS上のosquery／PowerShell等で実機確認",
        "カスタムSecurity Finding／Complianceテーブル",
        "ACC標準収集・標準CMDBマッピングとしての公式明記なし。",
    )


def status_fill(text: str) -> PatternFill:
    if text.startswith("○"):
        return PatternFill("solid", fgColor="E2F0D9")
    if text.startswith("△"):
        return PatternFill("solid", fgColor="FFF2CC")
    return PatternFill("solid", fgColor="FCE4D6")


def main() -> None:
    workbook = load_workbook(SOURCE)
    sheet = next(
        ws
        for ws in workbook.worksheets
        if ws.title.endswith("_JP") and not ws.title.endswith("(2)")
    )

    # Extend existing title and note merges to include the new comparison column.
    if "A1:L1" in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.unmerge_cells("A1:L1")
        sheet.merge_cells("A1:M1")
    if "A189:L191" in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.unmerge_cells("A189:L191")
        sheet.merge_cells("A189:M191")

    sheet["A1"] = (
        "MECM セキュリティ指摘 23 項目：取得フィールドと "
        "ServiceNow標準格納可否／ACC比較マトリクス"
    )

    # Reuse the existing visual style and add the ACC comparison header.
    source_header = sheet["L3"]
    target_header = sheet["M3"]
    target_header.value = "ACC取得・ServiceNow格納可否\n○ / △ / ×"
    target_header._style = copy(source_header._style)
    target_header.alignment = copy(source_header.alignment)
    target_header.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    sheet["M2"] = "○ 標準　△ カスタム・要確認　× ACC単独では不可"
    sheet["M2"].font = copy(sheet["L3"].font)
    sheet["M2"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    sheet["M2"].fill = PatternFill("solid", fgColor="D9EAF7")

    for row in range(4, 185):
        source_cell = sheet.cell(row=row, column=12)
        target_cell = sheet.cell(row=row, column=13)
        target_cell.value = assessment(str(sheet.cell(row=row, column=4).value or ""))
        target_cell._style = copy(source_cell._style)
        target_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        target_cell.fill = status_fill(target_cell.value)

    sheet.column_dimensions["M"].width = 49
    for row in range(4, 185):
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 15, 72)

    sheet["A189"] = (
        "判定基準：既存のMECM判定に加え、ACC比較列は "
        "○ = ServiceNow公式資料でACC-F／ACC-VCの標準収集および標準格納先が確認できる、"
        "△ = osquery・PowerShell・製品固有API等で取得可能性はあるが、カスタムCheck、変換、"
        "カスタムフィールド／Security Finding／Complianceテーブル等が必要、"
        "× = ACC単独では必要な元データまたは判定材料を取得できず、MECM、外部カタログ、"
        "顧客ライフサイクルマスタ等が必要、という意味である。"
        "ACCの『取得可能』は『標準CMDBへ自動格納』を意味しない。"
        "対象リリース、ACC-VCバージョン、プラグイン、SAM、権限、OSおよび実payloadで確認する。"
        "確認日：2026-06-25。\n"
        f"公式参考：{ACC_DATA_URL}\n{ACC_V_REFERENCE_URL}\n"
        f"{ACC_V_POLICY_URL}\n{ACC_CHECK_URL}"
    )
    sheet["A189"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    sheet.row_dimensions[189].height = 135

    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
