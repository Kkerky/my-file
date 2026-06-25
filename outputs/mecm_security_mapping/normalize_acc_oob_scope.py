from pathlib import Path
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


SOURCE = Path(os.environ["TARGET_XLSX"])
OUTPUT = SOURCE.with_name(SOURCE.stem + "_O-X判定修正版625.xlsx")


def main() -> None:
    workbook = load_workbook(SOURCE)
    sheet = next(
        ws
        for ws in workbook.worksheets
        if ws.title.endswith("_JP") and not ws.title.endswith("(2)")
    )

    sheet["M2"] = "○ 標準機能で取得・格納可能　× カスタマイズなしでは不可"
    sheet["M3"] = "ACC標準取得・ServiceNow格納可否\n○ / ×"

    converted = 0
    for row in range(4, 185):
        cell = sheet.cell(row=row, column=13)
        value = str(cell.value or "")
        if value.startswith("△"):
            details = value.splitlines()
            detail_text = "\n".join(details[1:])
            cell.value = (
                "× カスタマイズなしでは取得・格納不可\n"
                "理由：ACCの標準収集・標準CMDBマッピングとして確認できない。\n"
                f"参考（今回の判定範囲外）：{detail_text}"
            )
            cell.fill = PatternFill("solid", fgColor="FCE4D6")
            converted += 1

    # Clarify the SAM dependency wherever the standard software-install route is used.
    for row in range(4, 185):
        cell = sheet.cell(row=row, column=13)
        value = str(cell.value or "")
        if value.startswith("○") and "cmdb_sam_sw_install" in value:
            if "SAM" not in value.splitlines()[-1]:
                cell.value = (
                    value
                    + "\n前提：SAM／SAM Foundationが有効な場合は"
                    "cmdb_sam_sw_installを使用。未導入の場合は"
                    "cmdb_software_instance等を使用する。"
                )

    sheet["A189"] = (
        "今回のACC判定範囲：カスタマイズを行わず、ACC／ACC-VCの標準Check・Policyで取得し、"
        "ServiceNowの標準処理で標準テーブルへ格納できるものだけを○とする。"
        "osquery、PowerShell、WMI、独自Check、独自変換、カスタムフィールド、"
        "カスタムSecurity Finding／Complianceテーブル、外部マスタとの照合が必要な項目は、"
        "技術的に実現可能であっても本調査では×とする。"
        "cmdb_sam_sw_installはSoftware Installationテーブルであり、"
        "SAM／Software Asset Management Foundationの機能が有効な環境で使用される。"
        "SAM未導入時は、ACC-V／Discoveryのソフトウェア情報は"
        "cmdb_software_instance等に格納される場合がある。"
        "対象リリース、ACC-VC Content、SAMライセンス／プラグインおよび実payloadで確認する。"
        "確認日：2026-06-25。\n"
        "公式参考：\n"
        "https://www.servicenow.com/docs/r/it-operations-management/"
        "agent-client-collector/acc-visibility-checks-policies.html\n"
        "https://www.servicenow.com/docs/r/it-operations-management/"
        "agent-client-collector/using-enhanced-discovery-and-sam-together.html\n"
        "https://www.servicenow.com/docs/r/xanadu/it-service-management/"
        "software-asset-management-foundation-plugin/c_SAMDiscoverySAMF.html"
    )

    workbook.save(OUTPUT)
    print(f"converted={converted}")
    print(OUTPUT)


if __name__ == "__main__":
    main()
